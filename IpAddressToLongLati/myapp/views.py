from django.shortcuts import render
import openpyxl
from ip2geotools.databases.noncommercial import DbIpCity
import xlwt
from django.http import HttpResponse


def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["sheet 1"]

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="GeoLocation.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('GeoLocation')

        row_num = 0

        columns = ['IP Address', 'Longitude Value', 'Latitude Value']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num])

        for row in worksheet.iter_rows():
            row_num += 1
            for cell in range(len(row)):
                try:
                    valid_ip = DbIpCity.get(row[cell].value, api_key='free')
                    ws.write(row_num, 0, row[cell].value)
                    ws.write(row_num, 1, str(valid_ip.longitude))
                    ws.write(row_num, 2, str(valid_ip.latitude))

                except:
                    ws.write(row_num, 0, row[cell].value)

        wb.save(response)
        return response
